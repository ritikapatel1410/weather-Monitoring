#!/usr/bin/env python
# coding: utf-8

# In[ ]:


'''
 * File Description:- Python program to find current weather details of any city using openweathermap api
 * 
 * Name                 Date            Changes
 * Ritika Patidar       16/07/20        Initial Version
'''
# import required modules 
import requests, json 
import openpyxl
import threading
# Enter your API key here 
api_key = "a74fc0f54eaf493d76d96a9c7782ff53"

# base_url variable to store url 
base_url = "http://api.openweathermap.org/data/2.5/weather?"

# Give city name 
city_name = "Ujjain"

# complete_url variable to store 
# complete url address 
complete_url = base_url + "q=" + city_name + "&appid=" + api_key
#print(complete_url)

# get method of requests module 
# return response object 
def weather_monitor():
    threading.Timer(10.0, weather_monitor).start()
    response = requests.get(complete_url) 

    # json method of response object 
    # convert json format data into 
    # python format data 
    x = response.json() 
    #print(x)

    # Now x contains list of nested dictionaries 
    # Check the value of "cod" key is equal to 
    # "404", means city is found otherwise, 
    # city is not found 
    if x["cod"] != "404": 

        # store the value of "main" 
        # key in variable y 
        y = x["main"] 

        # store the value corresponding 
        # to the "temp" key of y 
        current_temperature = y["temp"]-273.15 

        # store the value corresponding 
        # to the "humidity" key of y 
        current_humidiy = y["humidity"] 

        # store the value of "weather" 
        # key in variable z 
        z = x["id"] 
        Z=x["sys"]["country"]
        # store the value corresponding 
        # to the "description" key at 
        # the 0th index of z 

        # print following values 
        print(" Temperature (in celsius unit) = " +
                        str(current_temperature) +
            "\n humidity (in percentage) = " +
                        str(current_humidiy) +
            "\n city token = " +
                        str(Z)+str(z)) 

        #add excel sheet path here
        wbkName ="C:/Users/pcs/Desktop/Project Sample.xlsx"
        wbk = openpyxl.load_workbook(wbkName)
        for wks in wbk.worksheets:
            wks.cell(row=11, column=1).value = str(Z)+str(z)
            wks.cell(row=11, column=2).value = current_temperature
            wks.cell(row=11, column=3).value = current_humidiy
            wks.cell(row=11, column=4).value = "C"
            wks.cell(row=11, column=5).value = 1
        wbk.save(wbkName)
        wbk.close
    else: 
        print(" City Not Found ") 

weather_monitor()


# In[49]:





# In[43]:





# In[ ]:




